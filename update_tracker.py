#!/usr/bin/env python3
"""
update_tracker.py — Build-Value Tracker

Walks a project's source files, counts raw + code LOC, applies a per-file-type
effort factor (to discount generated/repetitive content), groups files into named
modules, and writes:

  - loc_data.csv : Module, Raw LOC, Code LOC, Effort factor
  - (optional) the "Data" tab of Build_Value_Tracker.xlsx (--xlsx)

The xlsx "Tracker" tab does the maths live from the Data tab (equiv LOC →
engineer-days → hours → rand), with editable assumptions and a sensitivity band.

IMPORTANT FRAMING: this is a REPLACEMENT-COST estimate — what it would plausibly
cost to commission this build from a senior contractor. It is NOT revenue and NOT
realised value. Never quote the headline without the sensitivity band; the
LOC/day assumption moves the answer more than the rate does.

Usage:
  python3 update_tracker.py <folder> [--xlsx [path]] [--csv path]
                            [--rate 650] [--hours 8] [--locday 150]
                            [--exclude dir ...]

Examples:
  python3 update_tracker.py .
  python3 update_tracker.py . --xlsx Build_Value_Tracker.xlsx
"""
import argparse
import csv
import fnmatch
import os
import sys

# ── Source extensions we count ────────────────────────────────────────────────
EXTS = {'.html', '.js', '.py', '.md', '.css', '.svg', '.json', '.yaml', '.yml',
        '.sql', '.ts', '.tsx', '.jsx'}

# ── Effort factor per file type ───────────────────────────────────────────────
# Discount generated/repetitive content so the estimate survives a technical
# audience. App logic = 1.0; dense/markup/data files are worth less per line.
EFFORT = {
    '.svg': 0.30,   # mostly machine coordinates
    '.json': 0.30,  # data/config, not hand-reasoned logic
    '.html': 0.55,  # markup
    '.css': 0.60,   # styling, fairly repetitive
    '.md': 0.60,    # prose/docs
    '.yaml': 0.50, '.yml': 0.50,  # config
    '.sql': 0.90,   # hand-written queries
    # app logic — full weight
    '.js': 1.0, '.jsx': 1.0, '.ts': 1.0, '.tsx': 1.0, '.py': 1.0,
}
DEFAULT_EFFORT = 1.0

# ── Comment syntax (for stripping to "code" LOC) ──────────────────────────────
LINE_COMMENTS = {
    '.js': ['//'], '.jsx': ['//'], '.ts': ['//'], '.tsx': ['//'],
    '.py': ['#'], '.yaml': ['#'], '.yml': ['#'], '.sql': ['--'],
    '.css': [], '.md': [], '.html': [], '.svg': [], '.json': [],
}
C_BLOCK = {'.js', '.jsx', '.ts', '.tsx', '.css'}       # /* ... */
HTML_BLOCK = {'.html', '.svg', '.md'}                  # <!-- ... -->

# ── Module mapping ────────────────────────────────────────────────────────────
# Ordered (pattern, module) — FIRST match wins. Patterns are fnmatch globs tested
# (case-insensitively) against the file's path relative to the project root, with
# forward slashes. Edit freely. Versioned-filename patterns (e.g. *v5*, *v6b*) and
# *overview* are included so the map also works on versioned builds.
MODULE_RULES = [
    # Versioned / named filename patterns (take precedence, per request)
    ('*overview*',            'Docs — Product overview'),
    ('*v6b*',                 'Build — v6b'),
    ('*v6*',                  'Build — v6'),
    ('*v5*',                  'Build — v5'),

    # Docs
    ('docs/*pitch*',          'Docs — Experience OS pitch'),
    ('docs/specs/*',          'Docs — Specs'),
    ('docs/*',                'Docs'),
    ('*.md',                  'Docs'),

    # Tests
    ('test/*',                'Tests'),
    ('*test*',                'Tests'),

    # Client
    ('client/src/pages/*',            'Client — Pages'),
    ('client/src/os/*',               'Client — Inbox (OS)'),
    ('client/src/components/tiles/*', 'Client — Tiles'),
    ('client/src/components/editor/*','Client — Editor'),
    ('client/src/components/*',       'Client — Components'),
    ('client/src/lib/*',              'Client — Lib & shell'),
    ('client/src/*.css',              'Client — Styles'),
    ('client/src/*',                  'Client — App shell'),
    ('client/*',                      'Client — Config'),

    # Server (campaign/engagement engine grouped together)
    ('server/action*.js',          'Server — Campaign engine'),
    ('server/campaigntemplates.js','Server — Campaign engine'),
    ('server/segments.js',         'Server — Campaign engine'),
    ('server/billing.js',          'Server — Campaign engine'),
    ('server/scheduler.js',        'Server — Digests & scheduler'),
    ('server/onboarding.js',       'Server — Onboarding'),
    ('server/insights.js',         'Server — AI & insights'),
    ('server/mailer.js',           'Server — Messaging & email'),
    ('server/messaging.js',        'Server — Messaging & email'),
    ('server/push.js',             'Server — Messaging & email'),
    ('server/tileimg.js',          'Server — Messaging & email'),
    ('server/meta.js',             'Server — Ad-platform sync'),
    ('server/tiktok.js',           'Server — Ad-platform sync'),
    ('server/os.js',               'Server — Experience OS spine'),
    ('server/looker.js',           'Server — Data & Looker'),
    ('server/db.js',               'Server — Data & Looker'),
    ('server/store.js',            'Server — Data & Looker'),
    ('server/convert.js',          'Server — Data & Looker'),
    ('server/drill.js',            'Server — Data & Looker'),
    ('server/migrate.js',          'Server — Data & Looker'),
    ('server/recreate.js',         'Server — Data & Looker'),
    ('server/auth.js',             'Server — Core, auth & API'),
    ('server/roles.js',            'Server — Core, auth & API'),
    ('server/ratelimit.js',        'Server — Core, auth & API'),
    ('server/index.js',            'Server — Core, auth & API'),
    ('server/data/*',              'Server — Dashboard configs (data)'),
    ('server/fixtures/*',          'Server — Fixtures & data'),
    ('server/*',                   'Server — Other'),

    ('*',                          'Other'),
]

# ── Things we never count (deps / build output / generated) ───────────────────
EXCLUDE_DIRS = {'node_modules', '.git', 'dist', 'build', '.next', 'coverage',
                '.venv', 'venv', '__pycache__', '.cache', 'out'}
EXCLUDE_FILES = {'package-lock.json', 'yarn.lock', 'pnpm-lock.yaml',
                 'loc_data.csv'}


def module_for(relpath):
    p = relpath.replace(os.sep, '/').lower()
    for pattern, name in MODULE_RULES:
        if fnmatch.fnmatch(p, pattern):
            return name
    return 'Other'


def count_file(path, ext):
    """Return (raw_loc, code_loc). code_loc strips blank + comment lines."""
    try:
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()
    except OSError:
        return 0, 0
    prefixes = LINE_COMMENTS.get(ext, [])
    c_block = ext in C_BLOCK
    h_block = ext in HTML_BLOCK
    state = {'c': False, 'h': False}  # inside /* */ or <!-- -->
    raw = code = 0
    for line in lines:
        raw += 1
        s = line.strip()
        if not s:
            continue
        if _line_has_code(s, prefixes, c_block, h_block, state):
            code += 1
    return raw, code


def _line_has_code(s, prefixes, c_block, h_block, state):
    """Walk the line, honouring block/line comments, return True if any real
    (non-comment, non-blank) content remains."""
    i, n = 0, len(s)
    real = []
    while i < n:
        if state['c']:
            j = s.find('*/', i)
            if j == -1:
                i = n
            else:
                state['c'] = False
                i = j + 2
            continue
        if state['h']:
            j = s.find('-->', i)
            if j == -1:
                i = n
            else:
                state['h'] = False
                i = j + 3
            continue
        if prefixes and any(s.startswith(p, i) for p in prefixes):
            break  # line comment → rest is comment
        if c_block and s.startswith('/*', i):
            state['c'] = True
            i += 2
            continue
        if h_block and s.startswith('<!--', i):
            state['h'] = True
            i += 4
            continue
        ch = s[i]
        if not ch.isspace():
            real.append(ch)
        i += 1
    return len(real) > 0


def walk(folder, extra_excludes):
    excl = EXCLUDE_DIRS | set(extra_excludes or [])
    rows = []  # (relpath, ext, raw, code)
    for root, dirs, files in os.walk(folder):
        dirs[:] = [d for d in dirs if d not in excl]
        for fn in files:
            if fn in EXCLUDE_FILES:
                continue
            ext = os.path.splitext(fn)[1].lower()
            if ext not in EXTS:
                continue
            if fn.endswith(('.min.js', '.min.css')):
                continue
            full = os.path.join(root, fn)
            rel = os.path.relpath(full, folder)
            raw, code = count_file(full, ext)
            rows.append((rel, ext, raw, code))
    return rows


def aggregate(files):
    """Group files into modules. Per-module effort factor = code-LOC-weighted
    average of its files' factors, so (Code LOC × Effort factor) == equiv LOC."""
    mods = {}
    for rel, ext, raw, code in files:
        name = module_for(rel)
        ef = EFFORT.get(ext, DEFAULT_EFFORT)
        m = mods.setdefault(name, {'raw': 0, 'code': 0, 'equiv': 0.0})
        m['raw'] += raw
        m['code'] += code
        m['equiv'] += code * ef
    out = []
    for name, m in mods.items():
        factor = round(m['equiv'] / m['code'], 3) if m['code'] else round(
            sum(EFFORT.get(e, DEFAULT_EFFORT) for _, e, _, _ in files) / max(1, len(files)), 3)
        out.append({'module': name, 'raw': m['raw'], 'code': m['code'],
                    'factor': factor, 'equiv': round(m['equiv'])})
    out.sort(key=lambda r: r['equiv'], reverse=True)
    return out


def write_csv(rows, path):
    with open(path, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(['Module', 'Raw LOC', 'Code LOC', 'Effort factor'])
        for r in rows:
            w.writerow([r['module'], r['raw'], r['code'], r['factor']])


# ── Excel ─────────────────────────────────────────────────────────────────────
DATA_SHEET = 'Data'
TRACKER_SHEET = 'Tracker'
MODULE_ROWS = 100  # pre-built formula rows in Tracker (auto-blank past the data)


def write_xlsx(rows, path, rate, hours, locday, mult=2.0):
    try:
        import openpyxl
    except ImportError:
        print('  ! openpyxl not installed — skipping --xlsx (pip install openpyxl)',
              file=sys.stderr)
        return False
    from openpyxl import Workbook, load_workbook

    if os.path.exists(path):
        wb = load_workbook(path)  # preserve the Tracker tab + its formulas
    else:
        wb = Workbook()
        wb.remove(wb.active)
    # (Re)write the Data tab — plain table, no formulas.
    if DATA_SHEET in wb.sheetnames:
        del wb[DATA_SHEET]
    ws = wb.create_sheet(DATA_SHEET, 0)
    ws.append(['Module', 'Raw LOC', 'Code LOC', 'Effort factor'])
    for r in rows:
        ws.append([r['module'], r['raw'], r['code'], r['factor']])
    _style_data(ws)
    # Build the Tracker tab if it's missing (first run / fresh file).
    if TRACKER_SHEET not in wb.sheetnames:
        _build_tracker(wb, rate, hours, locday, mult)
    # Force Excel/Sheets to recalculate formulas on open (we write none cached).
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass
    wb.save(path)
    return True


def _style_data(ws):
    from openpyxl.styles import Font
    for c in ws[1]:
        c.font = Font(bold=True)
    widths = {'A': 32, 'B': 12, 'C': 12, 'D': 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        row[0].number_format = '0.000'


def _build_tracker(wb, rate, hours, locday, mult=2.0):
    from openpyxl.styles import Font, PatternFill, Alignment
    ws = wb.create_sheet(TRACKER_SHEET)
    yellow = PatternFill('solid', fgColor='FFF2CC')
    bold = Font(bold=True)
    big = Font(bold=True, size=14)
    muted = Font(color='808080', italic=True)
    wrap = Alignment(wrap_text=True, vertical='top')

    ws['A1'] = 'Build-Value Tracker — replacement-cost estimate'
    ws['A1'].font = big
    ws['A2'] = ('What it would plausibly cost to commission this build from a senior '
                'contractor at the rate below. This is a REPLACEMENT-COST estimate — '
                'NOT revenue, NOT realised value. "Build only" = writing the code; '
                '"Full delivery" applies the delivery multiplier for design, testing, '
                'QA, UAT and PM. Always read it with the sensitivity band: the Net '
                'LOC/day assumption moves the answer more than the rate, so quote the '
                'band, never a single headline figure.')
    ws['A2'].font = muted
    ws['A2'].alignment = wrap
    ws.merge_cells('A2:E2')
    ws.row_dimensions[2].height = 72

    # Assumptions (yellow, editable)
    ws['A4'] = 'Assumptions — edit these'
    ws['A4'].font = bold
    for r, (label, val, fmt) in enumerate([
        ('Hourly rate (R/hour)', rate, '"R"#,##0'),
        ('Hours per day', hours, '0'),
        ('Net LOC per day', locday, '0'),
        ('Delivery multiplier (design, test, QA, UAT, PM)', mult, '0.0"x"'),
    ], start=5):
        ws.cell(r, 1, label)
        c = ws.cell(r, 2, val)
        c.fill = yellow
        c.font = bold
        c.number_format = fmt
    # named refs used below
    RATE, HRS, LOCDAY, MULT = '$B$5', '$B$6', '$B$7', '$B$8'

    # Per-module table (formulas pull live from the Data tab)
    hdr = 11
    ws.cell(hdr - 1, 1, 'Per-module value').font = bold
    headers = ['Module', 'Equiv LOC', 'Engineer-days', 'Engineer-hours', 'Build value (R)']
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(hdr, ci, h)
        c.font = bold
    first = hdr + 1
    last = hdr + MODULE_ROWS
    for i in range(MODULE_ROWS):
        r = first + i
        dr = 2 + i  # corresponding Data row
        ws.cell(r, 1, f'=IF(Data!A{dr}="","",Data!A{dr})')
        ws.cell(r, 2, f'=IF(Data!C{dr}="","",Data!C{dr}*Data!D{dr})')
        ws.cell(r, 3, f'=IF(B{r}="","",B{r}/{LOCDAY})')
        ws.cell(r, 4, f'=IF(C{r}="","",C{r}*{HRS})')
        ws.cell(r, 5, f'=IF(D{r}="","",D{r}*{RATE})')
        ws.cell(r, 2).number_format = '#,##0'
        ws.cell(r, 3).number_format = '#,##0.0'
        ws.cell(r, 4).number_format = '#,##0'
        ws.cell(r, 5).number_format = '"R"#,##0'

    # TOTAL row
    tot = last + 1
    ws.cell(tot, 1, 'TOTAL').font = bold
    for ci, fmt in [(2, '#,##0'), (3, '#,##0.0'), (4, '#,##0'), (5, '"R"#,##0')]:
        col = chr(64 + ci)
        c = ws.cell(tot, ci, f'=SUM({col}{first}:{col}{last})')
        c.font = bold
        c.number_format = fmt
    equiv_total = f'$B${tot}'   # total equiv LOC cell
    build_total = f'$E${tot}'   # total build-only value cell

    # Headline (current assumptions): build-only vs full delivery.
    hh = tot + 2
    ws.cell(hh, 1, 'Headline — current assumptions').font = bold
    ws.cell(hh + 1, 1, 'Build only (writing the code)')
    cb = ws.cell(hh + 1, 2, f'={build_total}'); cb.number_format = '"R"#,##0'; cb.font = bold
    ws.cell(hh + 2, 1, 'Full delivery (× delivery multiplier)')
    cf = ws.cell(hh + 2, 2, f'={build_total}*{MULT}'); cf.number_format = '"R"#,##0'; cf.font = bold
    ws.cell(hh + 2, 3, '= design, testing, QA, UAT, PM on top of build').font = muted

    # Sensitivity band — total value at 100 / 150 / 250 net LOC/day, both scopes.
    s0 = hh + 5
    ws.cell(s0 - 1, 1, 'Sensitivity — total value by Net LOC/day (the dominant assumption)').font = bold
    for ci, h in enumerate(['Net LOC/day', 'Engineer-days', 'Build value (R)', 'Full delivery (R)'], start=1):
        ws.cell(s0, ci, h).font = bold
    for k, ld in enumerate([100, 150, 250], start=1):
        r = s0 + k
        ws.cell(r, 1, ld)
        ws.cell(r, 2, f'={equiv_total}/{ld}')
        ws.cell(r, 2).number_format = '#,##0.0'
        ws.cell(r, 3, f'={equiv_total}/{ld}*{HRS}*{RATE}')
        ws.cell(r, 3).number_format = '"R"#,##0'
        ws.cell(r, 4, f'={equiv_total}/{ld}*{HRS}*{RATE}*{MULT}')
        ws.cell(r, 4).number_format = '"R"#,##0'
        if ld == locday:
            ws.cell(r, 5, '← current').font = muted

    # How to update
    h0 = s0 + 6
    ws.cell(h0, 1, 'How to update').font = bold
    steps = [
        '1. Run:  python3 update_tracker.py /path/to/project --xlsx Build_Value_Tracker.xlsx',
        '2. The script rewrites the Data tab (Module, Raw LOC, Code LOC, Effort factor) and writes loc_data.csv.',
        '3. This Tracker tab recalculates automatically when the file is (re)opened.',
        '4. Google Sheets: import loc_data.csv into the Data tab (replace existing data); this tab recalculates the same way.',
        '5. Adjust the yellow cells (rate, hours/day, net LOC/day, delivery multiplier) at any time — everything below updates.',
        '6. Effort factors discount generated/repetitive content (SVG ~0.30, JSON ~0.30, HTML ~0.55, CSS/MD ~0.60, app logic 1.0),',
        '   so the figure is a defensible replacement cost, not a vanity line count.',
        '7. Delivery multiplier: "Build only" is just writing code; "Full delivery" multiplies it for design, testing, QA, UAT and PM.',
        '   Coding is ~40-50% of total project effort, so ~1.8x (lean) to ~2.5x (thorough); default 2.0x. This compounds uncertainty —',
        '   keep build-only as the defensible core and present full delivery as a range.',
    ]
    for k, s in enumerate(steps, start=1):
        ws.cell(h0 + k, 1, s)

    widths = {'A': 40, 'B': 16, 'C': 16, 'D': 18, 'E': 16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def main():
    ap = argparse.ArgumentParser(description='Build-Value Tracker — count LOC and estimate replacement cost.')
    ap.add_argument('folder', help='Project folder to scan')
    ap.add_argument('--csv', default='loc_data.csv', help='CSV output path (default loc_data.csv)')
    ap.add_argument('--xlsx', nargs='?', const='Build_Value_Tracker.xlsx', default=None,
                    help='Also write the Data tab of this workbook (default Build_Value_Tracker.xlsx)')
    ap.add_argument('--rate', type=float, default=650, help='Hourly rate (R) for a fresh workbook (default 650)')
    ap.add_argument('--hours', type=float, default=8, help='Hours/day for a fresh workbook (default 8)')
    ap.add_argument('--locday', type=float, default=150, help='Net LOC/day for a fresh workbook (default 150)')
    ap.add_argument('--multiplier', type=float, default=2.0, help='Delivery multiplier — design/test/QA/UAT/PM on top of build (default 2.0)')
    ap.add_argument('--exclude', nargs='*', default=[], help='Extra directory names to skip')
    args = ap.parse_args()

    folder = os.path.abspath(args.folder)
    if not os.path.isdir(folder):
        print(f'Not a folder: {folder}', file=sys.stderr)
        sys.exit(1)

    files = walk(folder, args.exclude)
    rows = aggregate(files)
    write_csv(rows, args.csv)

    total_raw = sum(r['raw'] for r in rows)
    total_code = sum(r['code'] for r in rows)
    total_equiv = sum(r['equiv'] for r in rows)

    print(f'Scanned {len(files)} source files in {folder}')
    print(f'  Raw LOC: {total_raw:,}   Code LOC: {total_code:,}   Equiv LOC: {total_equiv:,}')
    print(f'  -> {args.csv} ({len(rows)} modules)')

    if args.xlsx:
        if write_xlsx(rows, args.xlsx, args.rate, args.hours, args.locday, args.multiplier):
            print(f'  -> {args.xlsx} (Data tab written; Tracker recalculates on open)')

    # Headline + band (never a single figure). Build-only and full delivery
    # (× the delivery multiplier for design/testing/QA/UAT/PM).
    def build_at(ld):
        return total_equiv / ld * args.hours * args.rate
    print(f'\nReplacement-cost estimate (NOT revenue) at R{args.rate:g}/hour, {args.hours:g} h/day:')
    print(f'  {"Net LOC/day":<14}{"Build only":>16}{f"Full delivery (x{args.multiplier:g})":>20}')
    for ld in (100, 150, 250):
        marker = '  <- current' if ld == args.locday else ''
        b = build_at(ld)
        print(f'  @ {ld:<12}{("R" + format(b, ",.0f")):>16}{("R" + format(b * args.multiplier, ",.0f")):>20}{marker}')
    print('Quote the band, not a single number. Build-only is the defensible core;')
    print('full delivery adds design/testing/QA/UAT/PM and compounds uncertainty.')


if __name__ == '__main__':
    main()
